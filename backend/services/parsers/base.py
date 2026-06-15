import logging
from pathlib import Path
from typing import Iterator
from openpyxl import load_workbook

log = logging.getLogger(__name__)


def get_sheets_and_columns(file_path: Path) -> dict[str, list[str]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        first_row = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            v = cell.value
            first_row.append(str(v) if v is not None else "")
        result[name] = first_row
    wb.close()
    return result


def read_columns_at_row(
    file_path: Path, sheet_name: str, header_row: int,
) -> list[str]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = []
    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for idx, cell in enumerate(row, start=1):
            v = cell.value
            headers.append(
                str(v).strip() if v is not None else f"col_{idx}"
            )
    wb.close()
    return headers


def build_col_index(
    headers: list[str], mappings: dict[str, str],
) -> dict[str, int]:
    col_index = {}
    for sys_field, file_col in mappings.items():
        file_col = file_col.strip()
        for i, h in enumerate(headers):
            if h == file_col:
                col_index[sys_field] = i
                break
    return col_index


def iter_data_rows(
    file_path: Path, sheet_name: str, header_row: int,
) -> Iterator[tuple]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=header_row + 1):
        values = tuple(cell.value for cell in row)
        if any(values):
            yield values
    wb.close()


def count_data_rows(
    file_path: Path, sheet_name: str, header_row: int,
) -> int:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    total = 0
    for row in ws.iter_rows(min_row=header_row + 1):
        if any(cell.value for cell in row):
            total += 1
    wb.close()
    return total
