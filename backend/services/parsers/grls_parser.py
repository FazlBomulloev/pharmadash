import logging
import re
import zipfile
from pathlib import Path
from openpyxl import load_workbook
from backend.services.normalize import (
    normalize_mnn, normalize_str, parse_bool, parse_date,
)
from backend.services.parsers.base import (
    iter_data_rows, build_col_index, read_columns_at_row,
    get_sheets_and_columns,
)

_STATUS_PREFIX_RE = re.compile(r"^grls\d{4}-\d{2}-\d{2}-\d+-", re.IGNORECASE)


def _clean_status_from_filename(stem: str) -> str:
    cleaned = _STATUS_PREFIX_RE.sub("", stem)
    return cleaned.replace("_", " ").strip()

log = logging.getLogger(__name__)

GRLS_FIELDS = [
    "mnn", "tm", "ru_holder", "lf_full", "dosage",
    "jnvlp", "ru_number",
    "reg_date", "expire_date", "cancel_date",
    "status",
]

REQUIRED_GRLS_FIELDS = ["mnn", "jnvlp"]

STRING_FIELDS = {"tm", "ru_holder", "lf_full", "dosage", "ru_number", "status"}
BOOL_FIELDS = {"jnvlp"}
DATE_FIELDS = {"reg_date", "expire_date", "cancel_date"}


def parse_grls_rows(
    file_path: Path,
    sheet_name: str,
    header_row: int,
    mappings: dict[str, str],
    status_from_filename: str | None = None,
    max_rows: int | None = None,
) -> list[dict]:
    headers = read_columns_at_row(file_path, sheet_name, header_row)
    col_index = build_col_index(headers, mappings)

    rows = []
    count = 0
    for values in iter_data_rows(file_path, sheet_name, header_row):
        record = {}
        for field in GRLS_FIELDS:
            idx = col_index.get(field)
            raw = values[idx] if idx is not None and idx < len(values) else None

            if field == "mnn":
                record["mnn_raw"] = normalize_mnn(raw)
            elif field in STRING_FIELDS:
                record[field] = normalize_str(raw) or None
            elif field in BOOL_FIELDS:
                record[field] = parse_bool(raw)
            elif field in DATE_FIELDS:
                record[field] = parse_date(raw)

        if not record.get("status") and status_from_filename:
            record["status"] = status_from_filename

        if not record.get("mnn_raw"):
            continue
        if "jnvlp" not in record:
            record["jnvlp"] = False
        if not record.get("status"):
            continue

        rows.append(record)
        count += 1
        if max_rows and count >= max_rows:
            break

    log.info("GRLS: распарсено %d строк из %s", len(rows), file_path.name)
    return rows


def extract_grls_archive(
    archive_path: Path, dest_dir: Path,
) -> list[Path]:
    dest_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(archive_path, "r") as zf:
        zf.extractall(dest_dir)
    xlsx_files = sorted(dest_dir.glob("*.xlsx"))
    log.info("GRLS-архив: распакован, %d xlsx-файлов", len(xlsx_files))
    return xlsx_files


def parse_grls_archive(
    archive_dir: Path,
    sheet_name: str,
    header_row: int,
    mappings: dict[str, str],
) -> list[dict]:
    all_rows = []
    for xlsx in sorted(archive_dir.glob("*.xlsx")):
        status = _clean_status_from_filename(xlsx.stem)

        wb = load_workbook(xlsx, read_only=True, data_only=True)
        if sheet_name in wb.sheetnames:
            actual_sheet = sheet_name
        else:
            actual_sheet = wb.sheetnames[0]
        wb.close()

        rows = parse_grls_rows(
            xlsx, actual_sheet, header_row, mappings,
            status_from_filename=status,
        )
        all_rows.extend(rows)
        log.info(
            "GRLS-архив: файл %s (лист '%s') → %d строк (статус '%s')",
            xlsx.name, actual_sheet, len(rows), status,
        )
    log.info("GRLS-архив: всего %d строк", len(all_rows))
    return all_rows


def get_grls_reference_file(archive_dir: Path) -> Path | None:
    files = sorted(archive_dir.glob("*.xlsx"))
    return files[0] if files else None
