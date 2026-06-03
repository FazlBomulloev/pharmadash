import io
import json
import logging
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1B3A5C")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUB_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SUB_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
DATA_FONT = Font(size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_FILL = PatternFill("solid", fgColor="F5F7FA")
PCT_FMT = "0.0%"
USD_FMT = '#,##0'


def _auto_width(ws, min_w=8, max_w=18):
    for col_cells in ws.columns:
        length = min_w
        for cell in col_cells:
            if cell.value:
                length = max(length, min(len(str(cell.value)) + 2, max_w))
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = length


def _style_header(cell, level=0):
    if level == 0:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    else:
        cell.fill = SUB_HEADER_FILL
        cell.font = SUB_HEADER_FONT
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    cell.border = THIN_BORDER


def export_avp_xlsx(rows, years, regions, market_name) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"АВП {market_name}"

    y_labels = [str(y) for y in sorted(years)[-3:]]

    meta_headers = ["МНН", "ЛФ для АВП"]
    summary_headers = [
        f"Общее USD {y_labels[2]}",
        f"HOS USD {y_labels[2]}",
        f"RET USD {y_labels[2]}",
        "Пр* Общее", "Пр* HOS", "Пр* RET",
        "Прирост USD", "Прирост UN",
    ]

    total_usd_headers = [
        f"USD {y_labels[0]}",
        f"USD {y_labels[1]}",
        f"USD {y_labels[2]}",
    ]
    total_un_headers = [
        f"UN {y_labels[0]}",
        f"UN {y_labels[1]}",
        f"UN {y_labels[2]}",
    ]

    region_usd_headers = []
    for reg in regions:
        for sector in ["HOS", "RET"]:
            region_usd_headers.extend([
                f"Пр* {reg} {sector}",
                f"{reg} {sector} {y_labels[2]} USD",
                f"Доля {reg} {sector}",
                f"{reg} {sector} {y_labels[1]} USD",
                f"Прирост {reg} {sector}",
                f"{reg} {sector} {y_labels[0]} USD",
            ])

    all_headers = (
        meta_headers + summary_headers
        + total_usd_headers + total_un_headers
        + region_usd_headers
    )

    for col_idx, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _style_header(cell)

    ws.freeze_panes = "C2"

    for row_idx, avp in enumerate(rows, 2):
        region_usd = (
            json.loads(avp.region_usd_json)
            if avp.region_usd_json else {}
        )
        region_comp = (
            json.loads(avp.region_competitors_json)
            if avp.region_competitors_json else {}
        )

        data = [avp.mnn, avp.lf_avp]

        data.extend([
            avp.total_usd_y3, avp.hos_usd_y3,
            avp.ret_usd_y3,
            avp.competitors_total, avp.competitors_hos,
            avp.competitors_ret,
            avp.usd_growth, avp.un_growth,
        ])

        data.extend([
            avp.total_usd_y1, avp.total_usd_y2,
            avp.total_usd_y3,
        ])
        data.extend([
            avp.total_un_y1, avp.total_un_y2,
            avp.total_un_y3,
        ])

        for reg in regions:
            rd = region_usd.get(reg, {})
            rc = region_comp.get(reg, {})
            for sector_key in ["hos", "ret"]:
                sd = rd.get(sector_key, {})
                y3 = sd.get("y3", 0)
                y2 = sd.get("y2", 0)
                y1 = sd.get("y1", 0)
                total_y3 = avp.total_usd_y3 or 1
                share = y3 / total_y3 if total_y3 > 0 else 0
                growth = (
                    (y3 - y2) / y2 if y2 > 0 else None
                )
                comp = rc.get(sector_key, 0)
                data.extend([comp, y3, share, y2, growth, y1])

        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

            header = all_headers[col_idx - 1] if col_idx <= len(all_headers) else ""
            if "Прирост" in header or "Доля" in header:
                cell.number_format = PCT_FMT
            elif "USD" in header or "UN" in header:
                cell.number_format = USD_FMT

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log.info("Экспорт АВП: %d строк", len(rows))
    return buf


def export_kap_xlsx(rows, years, regions, market_name) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"КАП {market_name}"

    y_labels = [str(y) for y in sorted(years)[-3:]]

    headers = [
        "МНН", "НАШ АТХ", "ЛФ для АВП",
        f"Конк {y_labels[2]}", "ОСН. КОНК RET", "ОСН. КОНК",
    ]

    for y in y_labels:
        headers.extend([f"UN HOS {y}", f"UN RET {y}"])
    for y in y_labels:
        headers.extend([f"USD HOS {y}", f"USD RET {y}"])

    headers.extend(["Прирост UN", "Прирост USD"])
    headers.extend(["БГ", "Г", "ДОЛЯ БГ"])

    for reg in regions:
        headers.append(f"Доля {reg}")
    for reg in regions:
        headers.append(f"Конк {reg} {y_labels[1]}")
    for reg in regions:
        headers.append(f"Конк {reg} {y_labels[2]}")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _style_header(cell)

    ws.freeze_panes = "D2"

    for row_idx, kap in enumerate(rows, 2):
        region_shares = (
            json.loads(kap.region_shares_json)
            if kap.region_shares_json else {}
        )
        region_comp = (
            json.loads(kap.region_competitors_json)
            if kap.region_competitors_json else {}
        )

        data = [
            kap.mnn, kap.atc or "", kap.lf_avp,
            kap.competitors_count,
            kap.main_competitor_ret or "",
            kap.main_competitor_total or "",
        ]

        data.extend([
            kap.un_hos_y1, kap.un_ret_y1,
            kap.un_hos_y2, kap.un_ret_y2,
            kap.un_hos_y3, kap.un_ret_y3,
        ])
        data.extend([
            kap.usd_hos_y1, kap.usd_ret_y1,
            kap.usd_hos_y2, kap.usd_ret_y2,
            kap.usd_hos_y3, kap.usd_ret_y3,
        ])

        data.extend([kap.un_growth, kap.usd_growth])
        data.extend([kap.bg_count, kap.g_count, kap.bg_share])

        for reg in regions:
            data.append(region_shares.get(reg, 0))
        for reg in regions:
            rc = region_comp.get(reg, {})
            data.append(rc.get("y2", 0))
        for reg in regions:
            rc = region_comp.get(reg, {})
            data.append(rc.get("y3", 0))

        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

            header = headers[col_idx - 1] if col_idx <= len(headers) else ""
            if "Прирост" in header or "Доля" in header or "ДОЛЯ" in header:
                cell.number_format = PCT_FMT
            elif "USD" in header or "UN" in header:
                cell.number_format = USD_FMT

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log.info("Экспорт КАП: %d строк", len(rows))
    return buf
