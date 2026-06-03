import logging
import re
from pathlib import Path
from openpyxl import load_workbook

log = logging.getLogger(__name__)


def get_sheets_and_columns(
    file_path: Path,
) -> dict[str, list[str]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    result: dict[str, list[str]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        first_row = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            v = cell.value
            first_row.append(str(v) if v is not None else "")
        result[name] = first_row
    wb.close()
    return result


def read_columns_at_row(
    file_path: Path,
    sheet_name: str,
    header_row: int,
) -> list[str]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers: list[str] = []
    for row in ws.iter_rows(
        min_row=header_row, max_row=header_row
    ):
        for cell in row:
            v = cell.value
            if v is not None:
                headers.append(str(v).strip())
            else:
                headers.append(f"col_{cell.column}")
    wb.close()
    return headers


def _normalize_str(value) -> str:
    if value is None:
        return ""
    s = str(value).strip().upper()
    s = re.sub(r"\s+", " ", s)
    return s


def _to_float(value) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _compute_bg_g(tm: str, mnn: str) -> str:
    tm_norm = _normalize_str(tm)
    mnn_norm = _normalize_str(mnn)
    if not tm_norm or not mnn_norm:
        return "БГ"
    if tm_norm == mnn_norm:
        return "Г"
    if mnn_norm in tm_norm.split():
        return "Г"
    return "БГ"


SYSTEM_FIELDS = [
    "mnn", "tm", "producer", "sector", "region",
    "atc", "lf", "lf_avp", "strength", "pack_size",
    "country_mfr", "bg_g",
    "usd_y1", "usd_y2", "usd_y3",
    "un_y1", "un_y2", "un_y3",
]

REQUIRED_FIELDS = [
    "mnn", "tm", "producer", "sector", "region",
    "lf_avp", "usd_y1", "usd_y2", "usd_y3",
    "un_y1", "un_y2", "un_y3",
]

STRING_FIELDS = {
    "mnn", "tm", "producer", "sector", "region",
    "atc", "lf", "lf_avp", "strength", "pack_size",
    "country_mfr", "bg_g",
}

NUMERIC_FIELDS = {
    "usd_y1", "usd_y2", "usd_y3",
    "un_y1", "un_y2", "un_y3",
}


def parse_rows(
    file_path: Path,
    sheet_name: str,
    header_row: int,
    mappings: dict[str, str],
    max_rows: int | None = None,
) -> list[dict]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    headers: list[str] = []
    for row in ws.iter_rows(
        min_row=header_row, max_row=header_row
    ):
        for cell in row:
            v = cell.value
            headers.append(str(v).strip() if v is not None else "")

    col_index: dict[str, int] = {}
    for sys_field, file_col in mappings.items():
        file_col_stripped = file_col.strip()
        for i, h in enumerate(headers):
            if h == file_col_stripped:
                col_index[sys_field] = i
                break

    has_bg_g = "bg_g" in col_index
    data_start = header_row + 1
    rows: list[dict] = []
    count = 0

    for row in ws.iter_rows(min_row=data_start):
        values = [cell.value for cell in row]
        if not any(values):
            continue

        mnn_idx = col_index.get("mnn")
        if mnn_idx is None or mnn_idx >= len(values):
            continue
        if values[mnn_idx] is None:
            continue

        record: dict = {}
        skip = False
        for sys_field in SYSTEM_FIELDS:
            idx = col_index.get(sys_field)
            if idx is not None and idx < len(values):
                raw = values[idx]
            else:
                raw = None

            if sys_field in STRING_FIELDS:
                record[sys_field] = _normalize_str(raw)
            elif sys_field in NUMERIC_FIELDS:
                record[sys_field] = _to_float(raw)
            else:
                record[sys_field] = raw

        for rf in REQUIRED_FIELDS:
            v = record.get(rf)
            if v is None or v == "" or v == 0.0:
                if rf in NUMERIC_FIELDS:
                    pass
                else:
                    skip = True
                    break
        if skip:
            continue

        if not has_bg_g or not record.get("bg_g"):
            record["bg_g"] = _compute_bg_g(
                record.get("tm", ""), record.get("mnn", "")
            )

        rows.append(record)
        count += 1
        if max_rows and count >= max_rows:
            break

    wb.close()
    log.info("Распарсено %d строк из %s/%s", len(rows), sheet_name, file_path.name)
    return rows


def count_data_rows(
    file_path: Path,
    sheet_name: str,
    header_row: int,
) -> int:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    total = 0
    for row in ws.iter_rows(min_row=header_row + 1):
        if any(cell.value for cell in row):
            total += 1
    wb.close()
    return total
